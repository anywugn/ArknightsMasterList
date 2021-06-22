from openpyxl import load_workbook

file = 'aml-input.xlsx'
wb = load_workbook(filename=file)
ss = wb['skills']  # skill sheet

print('max rows = ' + str(ss.max_row))

maxRow = ss.max_row
for i in range(2, maxRow + 1):
    if ss['A' + str(i)].value is not None:
        print(ss['A' + str(i)].value)
        pass


